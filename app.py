import streamlit as st
import pandas as pd
import PyPDF2
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Alignment
import fitz  # PyMuPDF
import io
import time

# Page configuration
st.set_page_config(page_title="PDF Line-by-Line Excel Comparator", layout="wide")

st.title("📄 ↔ 📊 PDF Line-by-Line Excel Comparator")
st.markdown("""
**Line-by-Line Comparison:** PDF की हर line को Excel की हर row से compare करें
""")

# Function definitions
def extract_pdf_lines(pdf_file):
    """PDF से lines निकालो (line-by-line)"""
    pdf_lines = []
    
    try:
        # PyMuPDF use करो जो better lines देता है
        pdf_doc = fitz.open(stream=pdf_file.read(), filetype="pdf")
        
        for page_num in range(len(pdf_doc)):
            page = pdf_doc.load_page(page_num)
            page_text = page.get_text()
            
            # Split into lines और clean करो
            lines = page_text.split('\n')
            for line_num, line in enumerate(lines, 1):
                line_clean = line.strip()
                if line_clean:  # Only non-empty lines
                    pdf_lines.append({
                        'page': page_num + 1,
                        'line_num': line_num,
                        'original_line': line,
                        'clean_line': line_clean,
                        'lower_line': line_clean.lower()
                    })
        
        pdf_doc.close()
        
    except Exception as e:
        st.error(f"PDF extraction error: {e}")
        # Fallback to PyPDF2
        pdf_file.seek(0)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        
        line_counter = 1
        for page_num, page in enumerate(pdf_reader.pages, 1):
            page_text = page.extract_text()
            lines = page_text.split('\n')
            
            for line in lines:
                line_clean = line.strip()
                if line_clean:
                    pdf_lines.append({
                        'page': page_num,
                        'line_num': line_counter,
                        'original_line': line,
                        'clean_line': line_clean,
                        'lower_line': line_clean.lower()
                    })
                    line_counter += 1
    
    return pdf_lines

def create_highlighted_excel_line_compare(excel_file, comparison_results, pdf_lines_sample):
    """Create Excel file with line-by-line comparison results"""
    # Read the original Excel file
    if excel_file.name.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excel_file)
    else:
        # Convert .xls to .xlsx
        df = pd.read_excel(excel_file)
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)
    
    ws = wb.active
    
    # Define styles
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    red_font = Font(color="FF0000", bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="00AA00", bold=True)
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_font = Font(color="FF9900", bold=True)
    
    # Map column names to indices
    col_map = {}
    for idx, cell in enumerate(ws[1], 1):
        col_map[cell.value] = idx
    
    # Apply formatting based on comparison results
    for result in comparison_results:
        col_idx = col_map.get(result['excel_column'])
        if not col_idx:
            continue
            
        cell = ws.cell(row=result['excel_row'], column=col_idx)
        
        if result['match_status'] == '✅ Perfect Match':
            cell.fill = green_fill
            cell.font = green_font
            cell.value = f"✅ {result['excel_value']}"
            
        elif result['match_status'] == '⚠️ Partial Match':
            cell.fill = yellow_fill
            cell.font = yellow_font
            cell.value = f"⚠️ {result['excel_value']}"
            
        elif result['match_status'] == '❌ No Match':
            cell.fill = red_fill
            cell.font = red_font
            cell.value = f"❌ {result['excel_value']}"
            
        elif result['match_status'] == '⚪ Empty':
            cell.value = f"⚪ {result['excel_value']}"
    
    # Add comparison summary sheet
    wb.create_sheet("Line-by-Line Analysis")
    analysis_ws = wb["Line-by-Line Analysis"]
    
    # Title
    analysis_ws['A1'] = "📊 PDF LINE-BY-LINE COMPARISON REPORT"
    analysis_ws['A1'].font = Font(size=16, bold=True)
    
    # Summary statistics
    total_checks = len(comparison_results)
    perfect_matches = len([r for r in comparison_results if r['match_status'] == '✅ Perfect Match'])
    partial_matches = len([r for r in comparison_results if r['match_status'] == '⚠️ Partial Match'])
    no_matches = len([r for r in comparison_results if r['match_status'] == '❌ No Match'])
    
    analysis_ws['A3'] = "📈 COMPARISON STATISTICS"
    analysis_ws['A3'].font = Font(bold=True)
    
    analysis_ws['A5'] = f"Total Excel Cells Checked: {total_checks}"
    analysis_ws['A6'] = f"✅ Perfect Matches: {perfect_matches} ({(perfect_matches/total_checks*100):.1f}%)"
    analysis_ws['A7'] = f"⚠️ Partial Matches: {partial_matches} ({(partial_matches/total_checks*100):.1f}%)"
    analysis_ws['A8'] = f"❌ No Matches: {no_matches} ({(no_matches/total_checks*100):.1f}%)"
    
    accuracy = (perfect_matches + partial_matches * 0.5) / total_checks * 100
    analysis_ws['A10'] = f"🎯 OVERALL ACCURACY: {accuracy:.1f}%"
    analysis_ws['A10'].font = Font(color="FF0000", bold=True, size=14)
    
    # PDF Lines Sample
    analysis_ws['A12'] = "📄 PDF LINES SAMPLE (First 50 lines)"
    analysis_ws['A12'].font = Font(bold=True)
    
    row_num = 14
    for i, line in enumerate(pdf_lines_sample[:50], 1):
        analysis_ws.cell(row=row_num, column=1, value=f"Line {i}:")
        analysis_ws.cell(row=row_num, column=2, value=line['clean_line'][:100])
        row_num += 1
    
    # Detailed comparison results
    row_num += 2
    analysis_ws.cell(row=row_num, column=1, value="🔍 DETAILED COMPARISON RESULTS")
    analysis_ws.cell(row=row_num, column=1).font = Font(bold=True)
    
    headers = ['Excel Cell', 'Excel Value', 'Match Status', 'Matched PDF Line', 'Similarity %']
    for col_idx, header in enumerate(headers, 1):
        analysis_ws.cell(row=row_num + 2, column=col_idx, value=header)
        analysis_ws.cell(row=row_num + 2, column=col_idx).font = Font(bold=True)
    
    data_start_row = row_num + 3
    for idx, result in enumerate(comparison_results[:100]):  # Show first 100 results
        analysis_ws.cell(row=data_start_row + idx, column=1, value=result['excel_cell'])
        analysis_ws.cell(row=data_start_row + idx, column=2, value=str(result['excel_value'])[:50])
        analysis_ws.cell(row=data_start_row + idx, column=3, value=result['match_status'])
        
        if result['matched_pdf_line']:
            analysis_ws.cell(row=data_start_row + idx, column=4, value=result['matched_pdf_line'][:50])
        else:
            analysis_ws.cell(row=data_start_row + idx, column=4, value="No match")
        
        analysis_ws.cell(row=data_start_row + idx, column=5, value=f"{result['similarity_percent']:.1f}%")
    
    # Auto-adjust column widths
    for column in analysis_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        analysis_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# Initialize session state
if 'compare_done' not in st.session_state:
    st.session_state.compare_done = False
if 'comparison_results' not in st.session_state:
    st.session_state.comparison_results = None
if 'excel_df' not in st.session_state:
    st.session_state.excel_df = None
if 'pdf_lines' not in st.session_state:
    st.session_state.pdf_lines = []
if 'highlighted_excel' not in st.session_state:
    st.session_state.highlighted_excel = None

# File upload section
st.header("1️⃣ Upload Files")

col1, col2 = st.columns(2)

with col1:
    st.subheader("PDF File (with line separators)")
    st.caption("PDF जिसमें हर line अलग है")
    pdf_file = st.file_uploader("Choose PDF file", type=['pdf'], key="pdf_uploader")

with col2:
    st.subheader("Excel File (to check)")
    st.caption("Excel जिसमें हर row PDF की line से match करना है")
    excel_file = st.file_uploader("Choose Excel file", type=['xlsx', 'xls'], key="excel_uploader")

# Compare button (only show when both files uploaded)
if pdf_file and excel_file:
    st.header("2️⃣ Start Line-by-Line Comparison")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        compare_btn = st.button(
            "🔍 START LINE-BY-LINE COMPARISON", 
            type="primary", 
            use_container_width=True,
            help="Click to compare PDF lines with Excel rows"
        )
    
    if compare_btn:
        with st.spinner("Extracting lines from PDF..."):
            # Reset file pointers
            pdf_file.seek(0)
            excel_file.seek(0)
            
            # Extract PDF lines
            pdf_lines = extract_pdf_lines(pdf_file)
            st.session_state.pdf_lines = pdf_lines
            
            # Read Excel
            excel_df = pd.read_excel(excel_file)
            st.session_state.excel_df = excel_df
            
            if not pdf_lines:
                st.error("No lines found in PDF! Please check the PDF file.")
                st.stop()
            
            if excel_df.empty:
                st.error("Excel file is empty!")
                st.stop()
            
            # Show extracted info
            st.info(f"📊 **Extracted:** {len(pdf_lines)} lines from PDF, {excel_df.shape[0]} rows from Excel")
            
            # Perform line-by-line comparison
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            comparison_results = []
            total_cells = excel_df.size
            processed = 0
            
            # For each cell in Excel, find best matching line in PDF
            for col_idx, column in enumerate(excel_df.columns):
                for row_idx, cell_value in enumerate(excel_df[column]):
                    processed += 1
                    progress = processed / total_cells
                    progress_bar.progress(progress)
                    
                    excel_row = row_idx + 2
                    cell_str = str(cell_value) if not pd.isna(cell_value) else ""
                    
                    if processed % 10 == 0:
                        status_text.text(f"Comparing cell {processed}/{total_cells}...")
                    
                    if not cell_str.strip():
                        # Empty cell
                        comparison_results.append({
                            'excel_cell': f"{column}{excel_row}",
                            'excel_column': column,
                            'excel_row': excel_row,
                            'excel_value': cell_str,
                            'matched_pdf_line': None,
                            'similarity_percent': 0,
                            'match_status': '⚪ Empty'
                        })
                        continue
                    
                    # Find best matching line in PDF
                    best_match = None
                    best_similarity = 0
                    best_line_info = None
                    
                    cell_lower = cell_str.lower()
                    cell_words = set(cell_lower.split())
                    
                    for pdf_line in pdf_lines:
                        line_lower = pdf_line['lower_line']
                        line_words = set(line_lower.split())
                        
                        if cell_lower == line_lower:
                            # Exact match found
                            best_match = pdf_line['clean_line']
                            best_similarity = 100
                            best_line_info = pdf_line
                            break
                        
                        # Calculate word similarity
                        if cell_words and line_words:
                            common_words = cell_words.intersection(line_words)
                            similarity = len(common_words) / len(cell_words) * 100
                            
                            if similarity > best_similarity:
                                best_similarity = similarity
                                best_match = pdf_line['clean_line']
                                best_line_info = pdf_line
                    
                    # Determine match status
                    if best_similarity == 100:
                        match_status = '✅ Perfect Match'
                    elif best_similarity >= 70:
                        match_status = '⚠️ Partial Match'
                    else:
                        match_status = '❌ No Match'
                    
                    comparison_results.append({
                        'excel_cell': f"{column}{excel_row}",
                        'excel_column': column,
                        'excel_row': excel_row,
                        'excel_value': cell_str,
                        'matched_pdf_line': best_match,
                        'pdf_line_info': best_line_info,
                        'similarity_percent': best_similarity,
                        'match_status': match_status
                    })
            
            st.session_state.comparison_results = comparison_results
            status_text.text("✅ Line-by-line comparison complete!")
            time.sleep(0.5)
            
            # Create highlighted Excel
            excel_file.seek(0)
            highlighted_excel = create_highlighted_excel_line_compare(
                excel_file, 
                comparison_results, 
                pdf_lines[:50]  # Send first 50 lines for sample
            )
            st.session_state.highlighted_excel = highlighted_excel
            
            st.session_state.compare_done = True
            
            # Show completion message
            st.success("🎯 Line-by-line comparison completed! View results below.")

# Show side-by-side comparison if comparison is done
if st.session_state.compare_done and st.session_state.comparison_results:
    st.header("3️⃣ Line-by-Line Comparison Results")
    
    # Statistics
    results_df = pd.DataFrame(st.session_state.comparison_results)
    total_checks = len(results_df)
    perfect_matches = len(results_df[results_df['match_status'] == '✅ Perfect Match'])
    partial_matches = len(results_df[results_df['match_status'] == '⚠️ Partial Match'])
    no_matches = len(results_df[results_df['match_status'] == '❌ No Match'])
    empty_cells = len(results_df[results_df['match_status'] == '⚪ Empty'])
    
    accuracy = (perfect_matches + partial_matches * 0.5) / total_checks * 100 if total_checks > 0 else 0
    
    # Display metrics
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("📊 Total Cells", total_checks)
    col2.metric("✅ Perfect", perfect_matches, f"{(perfect_matches/total_checks*100):.1f}%")
    col3.metric("⚠️ Partial", partial_matches, f"{(partial_matches/total_checks*100):.1f}%")
    col4.metric("❌ No Match", no_matches, f"-{(no_matches/total_checks*100):.1f}%")
    
    # Accuracy score
    st.subheader(f"🎯 Overall Accuracy: {accuracy:.1f}%")
    st.progress(accuracy / 100)
    st.success(f"✅ STATUS: COMPLETED - {total_checks} cells checked line-by-line")
    
    # Side-by-side layout
    st.subheader("📄 PDF Lines vs 📊 Excel Data")
    
    # Create two columns for side-by-side view
    left_col, right_col = st.columns(2)
    
    with left_col:
        st.markdown("### 📄 PDF Lines (First 30 lines)")
        
        pdf_lines_display = []
        for i, line in enumerate(st.session_state.pdf_lines[:30], 1):
            pdf_lines_display.append(f"Line {i}: {line['clean_line'][:80]}{'...' if len(line['clean_line']) > 80 else ''}")
        
        st.text_area("PDF Lines", "\n".join(pdf_lines_display), height=400, label_visibility='collapsed')
        
        st.caption(f"Total {len(st.session_state.pdf_lines)} lines extracted from PDF")
    
    with right_col:
        st.markdown("### 📊 Excel Data with Line Matches")
        
        # Create display dataframe
        display_df = st.session_state.excel_df.copy()
        
        # Add match indicators
        for result in st.session_state.comparison_results:
            if result['match_status'] == '✅ Perfect Match':
                indicator = "✅"
            elif result['match_status'] == '⚠️ Partial Match':
                indicator = "⚠️"
            elif result['match_status'] == '❌ No Match':
                indicator = "❌"
            elif result['match_status'] == '⚪ Empty':
                indicator = "⚪"
            else:
                indicator = ""
            
            col = result['excel_column']
            row_idx = result['excel_row'] - 2  # Convert to 0-based index
            
            if 0 <= row_idx < len(display_df):
                current_value = display_df.at[row_idx, col]
                if pd.notna(current_value):
                    display_df.at[row_idx, col] = f"{indicator} {current_value}"
        
        st.dataframe(
            display_df,
            use_container_width=True,
            height=400
        )
    
    # Detailed analysis
    st.subheader("🔍 Detailed Line-by-Line Analysis")
    
    # Tabs for different views
    tab1, tab2, tab3, tab4 = st.tabs(["All Results", "No Matches", "Perfect Matches", "PDF Line Mapping"])
    
    with tab1:
        display_results = results_df[['excel_cell', 'excel_value', 'match_status', 'similarity_percent']].copy()
        display_results['similarity_percent'] = display_results['similarity_percent'].apply(lambda x: f"{x:.1f}%")
        st.dataframe(display_results, use_container_width=True, height=300)
    
    with tab2:
        no_match_df = results_df[results_df['match_status'] == '❌ No Match']
        if not no_match_df.empty:
            st.warning(f"Found {len(no_match_df)} cells with no matching PDF line:")
            display_cols = ['excel_cell', 'excel_value', 'similarity_percent']
            no_match_df = no_match_df.copy()
            no_match_df['similarity_percent'] = no_match_df['similarity_percent'].apply(lambda x: f"{x:.1f}%")
            st.dataframe(no_match_df[display_cols], use_container_width=True, height=300)
        else:
            st.success("🎉 All cells have at least partial match with PDF lines!")
    
    with tab3:
        perfect_match_df = results_df[results_df['match_status'] == '✅ Perfect Match']
        if not perfect_match_df.empty:
            st.info(f"Found {len(perfect_match_df)} cells with perfect line matches:")
            
            # Show with PDF line mapping
            perfect_display = perfect_match_df[['excel_cell', 'excel_value', 'matched_pdf_line']].copy()
            perfect_display['matched_pdf_line'] = perfect_display['matched_pdf_line'].apply(
                lambda x: str(x)[:60] + "..." if x and len(str(x)) > 60 else str(x)
            )
            st.dataframe(perfect_display, use_container_width=True, height=300)
        else:
            st.warning("No perfect matches found")
    
    with tab4:
        st.write("**PDF Line to Excel Cell Mapping**")
        
        # Group by PDF line
        line_mapping = {}
        for result in st.session_state.comparison_results:
            if result['pdf_line_info']:
                line_text = result['pdf_line_info']['clean_line'][:50]
                if line_text not in line_mapping:
                    line_mapping[line_text] = []
                line_mapping[line_text].append({
                    'excel_cell': result['excel_cell'],
                    'similarity': result['similarity_percent']
                })
        
        # Display mapping
        for line_text, matches in list(line_mapping.items())[:20]:  # Show first 20
            with st.expander(f"📄 '{line_text}...'"):
                for match in matches:
                    st.write(f"- **{match['excel_cell']}**: {match['similarity']:.1f}% match")
    
    # Download section
    st.header("4️⃣ Download Results")
    
    if st.session_state.highlighted_excel:
        col1, col2 = st.columns([1, 2])
        
        with col1:
            st.download_button(
                label="📥 Download Highlighted Excel",
                data=st.session_state.highlighted_excel,
                file_name=f"LINE_COMPARE_{excel_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        
        with col2:
            st.markdown("""
            **Downloaded file contains:**
            - ✅ **Green cells** = Perfect line match (100%)
            - ⚠️ **Yellow cells** = Partial line match (70-99%)
            - ❌ **Red cells** = No line match (<70%)
            - ⚪ **Gray cells** = Empty cells
            - 📝 **Analysis sheet** = Line-by-line mapping details
            """)
    
    # Reset button
    st.markdown("---")
    if st.button("🔄 Start New Line Comparison", type="secondary"):
        st.session_state.compare_done = False
        st.session_state.comparison_results = None
        st.session_state.excel_df = None
        st.session_state.pdf_lines = []
        st.session_state.highlighted_excel = None
        st.rerun()

elif pdf_file and excel_file and not st.session_state.compare_done:
    st.info("👆 Click the 'START LINE-BY-LINE COMPARISON' button above to begin")

else:
    st.info("📁 Please upload both PDF and Excel files to start line-by-line comparison")

# Footer
st.markdown("---")
st.caption("🔧 Tool: PDF Line-by-Line Excel Comparator | Status: Ready")
